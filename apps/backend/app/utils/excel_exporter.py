"""Utility for Excel export.

Generates formatted Excel reports for projects.
"""

from __future__ import annotations

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

class ExcelExporter:
    @staticmethod
    def generate_project_report(project_data: dict) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Project Summary"
        
        # Header
        ws['A1'] = "NEXUS V3 — Project Report"
        ws['A1'].font = Font(size=14, bold=True)
        
        ws['A3'] = "Project Name:"
        ws['B3'] = project_data.get('name', 'N/A')
        
        ws['A4'] = "Status:"
        ws['B4'] = project_data.get('status', 'N/A')
        
        # Add some mock data sections
        ws['A6'] = "KPI"
        ws['B6'] = "Value"
        ws['A6'].font = Font(bold=True)
        ws['B6'].font = Font(bold=True)
        
        ws['A7'] = "SPI"
        ws['B7'] = project_data.get('spi', 1.0)
        
        ws['A8'] = "CPI"
        ws['B8'] = project_data.get('cpi', 1.0)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
