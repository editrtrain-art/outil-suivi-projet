from __future__ import annotations

import io
import uuid
from datetime import date
from typing import List
import structlog

# PDF imports
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# Excel imports
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from sqlalchemy import select
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from app.domain.models.project import Project
from app.domain.models.phase import Phase
from app.domain.models.task import Task
from app.domain.services.evm_calculator import EVMCalculator
from app.core.exceptions import NotFoundError

logger: structlog.stdlib.BoundLogger = structlog.get_logger(__name__)


class ExportService:
    """Service dedicated to compiling project performance exports in PDF and Excel formats."""

    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    async def generate_project_pdf(self, project_id: uuid.UUID, evm_data: dict) -> bytes:
        """Compile a professional project status summary report PDF using ReportLab.

        Args:
            project_id: UUID of the project.
            evm_data: Computed EVM metrics dictionary.

        Returns:
            bytes: Binary content of the compiled PDF report.
        """
        project = await self.db.get(Project, project_id)
        if not project:
            raise NotFoundError("Project", str(project_id))

        result = await self.db.execute(
            select(Task)
            .join(Phase, Task.phase_id == Phase.id)
            .where(Phase.project_id == project_id)
            .order_by(Task.wbs_code)
        )
        tasks = result.scalars().all()

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            name="TitleStyle",
            parent=styles["Heading1"],
            fontSize=22,
            leading=26,
            textColor=colors.HexColor("#1e3a8a"),
            spaceAfter=15,
        )
        h2_style = ParagraphStyle(
            name="Heading2Style",
            parent=styles["Heading2"],
            fontSize=14,
            leading=18,
            textColor=colors.HexColor("#0f172a"),
            spaceBefore=12,
            spaceAfter=8,
        )
        body_style = styles["Normal"]

        story = []

        # 1. Header
        story.append(Paragraph(f"NEXUS V3 — Project Status Report", title_style))
        story.append(Paragraph(f"Project: <b>{project.name}</b>", body_style))
        story.append(Paragraph(f"Status: {project.status.upper()} | Date: {date.today()}", body_style))
        story.append(Spacer(1, 15))

        # 2. Project metadata table
        meta_data = [
            ["Start Date", str(project.start_date), "End Date", str(project.end_date)],
            ["Budget (BAC)", f"{project.budget_total:,.2f} DH", "Active Baseline", "Initial Baseline"],
        ]
        meta_table = Table(meta_data, colWidths=[100, 150, 100, 150])
        meta_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#334155")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
                ("PADDING", (0, 0), (-1, -1), 6),
            ])
        )
        story.append(meta_table)
        story.append(Spacer(1, 15))

        # 3. EVM performance indices
        story.append(Paragraph("Earned Value Analysis", h2_style))
        evm_data_table = [
            ["Planned Value (PV)", f"{evm_data.get('pv', 0.0):,.2f} DH", "Schedule Variance (SV)", f"{evm_data.get('sv', 0.0):,.2f} DH"],
            ["Earned Value (EV)", f"{evm_data.get('ev', 0.0):,.2f} DH", "Cost Variance (CV)", f"{evm_data.get('cv', 0.0):,.2f} DH"],
            ["Actual Cost (AC)", f"{evm_data.get('ac', 0.0):,.2f} DH", "Schedule Index (SPI)", f"{evm_data.get('spi', 1.0):.2f}"],
            ["Estimate at Completion (EAC)", f"{evm_data.get('eac', 0.0):,.2f} DH", "Cost Index (CPI)", f"{evm_data.get('cpi', 1.0):.2f}"],
        ]
        evm_table = Table(evm_data_table, colWidths=[180, 100, 160, 100])
        evm_table.setStyle(
            TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
                ("BACKGROUND", (0, 0), (1, -1), colors.HexColor("#f1f5f9")),
                ("PADDING", (0, 0), (-1, -1), 6),
            ])
        )
        story.append(evm_table)
        story.append(Spacer(1, 15))

        # 4. Tasks schedules
        story.append(Paragraph("WBS Task Schedules", h2_style))
        task_headers = ["WBS", "Task Name", "Start", "Finish", "Float", "Critical"]
        task_rows = [task_headers]
        for t in tasks:
            task_rows.append([
                t.wbs_code,
                t.name[:25],
                str(t.start_date_scheduled) if t.start_date_scheduled else "-",
                str(t.end_date_scheduled) if t.end_date_scheduled else "-",
                f"{t.total_float}d",
                "YES" if t.is_critical else "NO",
            ])

        task_table = Table(task_rows, colWidths=[40, 180, 80, 80, 40, 50])
        task_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("PADDING", (0, 0), (-1, -1), 4),
                ("ALIGN", (4, 0), (-1, -1), "CENTER"),
            ])
        )
        story.append(task_table)

        doc.build(story)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes

    async def generate_project_excel(self, project_id: uuid.UUID) -> bytes:
        """Compile a structured Excel planning sheet containing WBS codes, dates, and float variables.

        Args:
            project_id: UUID of the project.

        Returns:
            bytes: Binary content of the generated Excel spreadsheet.
        """
        project = await self.db.get(Project, project_id)
        if not project:
            raise NotFoundError("Project", str(project_id))

        result = await self.db.execute(
            select(Task)
            .join(Phase, Task.phase_id == Phase.id)
            .where(Phase.project_id == project_id)
            .order_by(Task.wbs_code)
        )
        tasks = result.scalars().all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "WBS Tasks"

        # Headers
        headers = [
            "WBS Code",
            "Task Name",
            "Duration (Days)",
            "Start Date",
            "Finish Date",
            "Total Float",
            "Free Float",
            "Is Critical",
            "Status",
        ]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        border_style = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border_style

        # Add data
        for row_num, t in enumerate(tasks, 2):
            ws.append([
                t.wbs_code,
                t.name,
                t.duration_days,
                str(t.start_date_scheduled) if t.start_date_scheduled else "-",
                str(t.end_date_scheduled) if t.end_date_scheduled else "-",
                t.total_float,
                t.free_float,
                "YES" if t.is_critical else "NO",
                t.status.upper(),
            ])

            # Apply borders and alignments to data rows
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = border_style
                if col_num in [1, 3, 4, 5, 6, 7, 8, 9]:
                    cell.alignment = center_align

        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        return excel_bytes
